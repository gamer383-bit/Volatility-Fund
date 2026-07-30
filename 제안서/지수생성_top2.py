# -*- coding: utf-8 -*-
"""TOP2 지수 생성 (사용자 규칙, 조건 11)
- 기준 엑셀(삼성전자_하이닉스 시트)에서 삼성전자·SK하이닉스를 50%:50%로
  '매일 리밸런싱'(비용 없음)한 지수를 생성 → 시트 D열에 기록
- 지수_t = 지수_{t-1} × (1 + 0.5·r_삼성 + 0.5·r_하이닉스), 가장 오래된 날짜=100
- 행 정렬(오름/내림차순)과 무관하게 동작: 날짜로 정렬해 계산 후 원래 행 위치에 기록
"""
import sys, io, os, shutil, datetime
sys.stdout=io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')
import openpyxl

BASE=os.path.dirname(os.path.abspath(__file__))
XLS=os.path.join(os.path.dirname(BASE),'data','삼성전자_하이닉스_코스피_코스피200_코스닥150_최근5년_주가데이터.xlsx')
BK=XLS.replace('.xlsx',f"_backup_{datetime.date.today().strftime('%Y%m%d')}.xlsx")
if not os.path.exists(BK):
    shutil.copy2(XLS,BK); print("백업:",os.path.basename(BK))

wb=openpyxl.load_workbook(XLS)
ws=wb['삼성전자_하이닉스']
hdr={9:'TOP2',10:'TOP2지수(삼성·하이닉스 50:50 일별리밸런싱)',11:'-',12:'-',
     13:'지수(최초일=100, 리밸비용 0)',14:'일간'}
for r,v in hdr.items(): ws.cell(row=r,column=4,value=v)

rows=[]
for r in range(15,ws.max_row+1):
    dt=ws.cell(row=r,column=1).value
    sv=ws.cell(row=r,column=2).value; hv=ws.cell(row=r,column=3).value
    if dt is None or sv is None or hv is None: continue
    rows.append((dt,r,float(sv),float(hv)))
rows.sort(key=lambda t:t[0])                     # 날짜 오름차순으로 계산
idx=100.0; prev=None
for dt,r,sv,hv in rows:
    if prev is not None:
        idx*=1.0+0.5*(sv/prev[0]-1.0)+0.5*(hv/prev[1]-1.0)
    ws.cell(row=r,column=4,value=round(idx,6))
    prev=(sv,hv)
print(f"지수 기록: {len(rows)}행 · {rows[0][0].date()}(=100) ~ {rows[-1][0].date()} 최종 {idx:.2f}")
try:
    wb.save(XLS); print("[OK] D열 저장 완료")
except PermissionError:
    print("[잠김] 엑셀에서 파일을 닫아주세요")
